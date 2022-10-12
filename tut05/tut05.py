from openpyxl import load_workbook
import pandas as pd
import numpy as np

wb=load_workbook("octant_input.xlsx")
df=pd.read_excel("octant_input.xlsx")
ws=wb.worksheets[0]

#creating lists to hold values of u,v,w excluding 'u','v','w' using pandas
list1=df.U
list2=df.V
list3=df.W

#calculating the summation of u,v,w values and then dividing by number of elements to get averages
totalu=0
for x in range(0,len(list1)):
    totalu=totalu+list1[x]

Uavg=totalu/len(list1)

totalv=0
for x in range(0,len(list2)):
    totalv=totalv+list2[x]

Vavg=totalv/len(list2)

totalw=0
for x in range(0,len(list3)):
    totalw=totalw+list3[x]
    
Wavg=totalw/len(list3)

#forming lists of title and value for averages , and writing them to excel file
Ua=["Uavg",Uavg]
Va=["Vavg",Vavg]
Wa=["Wavg",Wavg]
for i in range(1,3):
    ws.cell(row=i,column=5).value=Ua[i-1]

for i in range(1,3):
    ws.cell(row=i,column=6).value=Va[i-1]
    
for i in range(1,3):
    ws.cell(row=i,column=7).value=Wa[i-1]

#list4,list5,list6 to store values of u-uavg , v-vavg , w-wavg as lists
list4=[]
for i in list1:
    tmp=i-Uavg
    list4.append(tmp)
    
ws.cell(row=1,column=8).value="U'"
for i in range(2,len(list4)+2):
    ws.cell(row=i,column=8).value=list4[i-2]
    
list5=[]
for i in list2:
    tmp=i-Vavg
    list5.append(tmp)
    
ws.cell(row=1,column=9).value="V'"
for i in range(2,len(list5)+2):
    ws.cell(row=i,column=9).value=list5[i-2]
    
list6=[]
for i in list3:
    tmp=i-Wavg
    list6.append(tmp)
    
ws.cell(row=1,column=10).value="W'"
for i in range(2,len(list6)+2):
    ws.cell(row=i,column=10).value=list6[i-2]

#list7 is created to hold the octant value for u',v',w' data, as a list
ws.cell(row=1,column=11).value="Octants"
list7=[]
for i in range (0,len(list1)):
    if list4[i]>0 and list5[i]>0:
        if list6[i]>0:
            list7.append(+1)
        else:
            list7.append(-1)
        
    if list4[i]>0 and list5[i]<0:
        if list6[i]>0:
            list7.append(+4)
        else:
            list7.append(-4)
        
    if list4[i]<0 and list5[i]>0:
        if list6[i]>0:
            list7.append(+2)
        else:
            list7.append(-2)
        
    if list4[i]<0 and list5[i]<0:
        if list6[i]>0:
            list7.append(+3)
        else:
            list7.append(-3)
        

for i in range(2,len(list7)+2):
    ws.cell(row=i,column=11).value=list7[i-2]
    
#creating the table for overall counts and adding the overall counts from list7
ws.cell(row=2,column=13).value="Overall Count"
ws.cell(row=3,column=12).value="User Input"
list8=[+1,-1,+2,-2,+3,-3,+4,-4]
for i in range(0,8):
    ws.cell(row=1,column=i+14).value=list8[i]
   
ws.cell(row=2,column=14).value=list7.count(+1)
ws.cell(row=2,column=15).value=list7.count(-1)
ws.cell(row=2,column=16).value=list7.count(+2)
ws.cell(row=2,column=17).value=list7.count(-2)
ws.cell(row=2,column=18).value=list7.count(+3)
ws.cell(row=2,column=19).value=list7.count(-3)
ws.cell(row=2,column=20).value=list7.count(+4)
ws.cell(row=2,column=21).value=list7.count(-4)

#giving value to mod , which can be changed and program will behave accordingly
mod=5000
ws.cell(row=3,column=13).value="Mod"+str(mod)

#creating a partition variable p
mod_ranges=[]
p=(len(list1)//mod)+1
cl=[] #cl holds empty lists based on number of partitions
cp=[] #cp holds empty lists based on number of partitions
for i in range(0,p):
    l=[]
    cl.append(l)
    m=[]
    cp.append(m)
    
a=0
#the empty lists in cl hold values of octants in the ranges which depend on value of mod or p
for y in range(0,p):
    for x in range(a,a+mod):
        if x<=len(list1)-1:
            cl[y].append(list7[x])
    a=a+mod

#writing the individual count in the table of variable dimensions , that depend on mod value or p value
#Created a nested list cp in which each list contains the counts of octants in mod ranges
for i in range(0,p): 
    if mod*(i+1)<=len(list1):
        ws.cell(row=4+i,column=13).value=str(mod*i)+"-"+str(mod*(i+1)-1)
    else:
        ws.cell(row=4+i,column=13).value=str(mod*i)+"-"+str(len(list1)-1)
    ws.cell(row=4+i,column=14).value=cl[i].count(+1)
    cp[i].append(cl[i].count(+1))
    ws.cell(row=4+i,column=15).value=cl[i].count(-1)
    cp[i].append(cl[i].count(-1))
    ws.cell(row=4+i,column=16).value=cl[i].count(+2)
    cp[i].append(cl[i].count(+2))
    ws.cell(row=4+i,column=17).value=cl[i].count(-2)
    cp[i].append(cl[i].count(-2))
    ws.cell(row=4+i,column=18).value=cl[i].count(+3)
    cp[i].append(cl[i].count(+3))
    ws.cell(row=4+i,column=19).value=cl[i].count(-3)
    cp[i].append(cl[i].count(-3))
    ws.cell(row=4+i,column=20).value=cl[i].count(+4)
    cp[i].append(cl[i].count(+4))
    ws.cell(row=4+i,column=21).value=cl[i].count(-4)
    cp[i].append(cl[i].count(-4))
    

#defining 3 lists to write in excel as a for loop
rank_title=["Rank Octant 1","Rank Octant -1","Rank Octant 2","Rank Octant -2","Rank Octant 3","Rank Octant -3","Rank Octant 4","Rank Octant -4"]
rank_title2=["Rank 1 Octant ID","Rank 1 Octant Name"]
octant_names=["Internal outward interaction","External Outward Interaction","External Ejection","Internal Ejection","External Inward Interaction","Internal Inward Interaction","Internal Sweep","External Sweep"]

#creating the skeleton in excel 
for i in range(0,2):
    ws.cell(row=2,column=30+i).value=rank_title2[i]
for i in range(0,8):
    ws.cell(row=2,column=22+i).value=rank_title[i]

ws.cell(row=8+p,column=14).value="Octant ID"
ws.cell(row=8+p,column=15).value="Octant Name"
ws.cell(row=8+p,column=16).value="Count Of Rank 1 Mod Values"
for i in range(0,8):
    ws.cell(row=9+p+i,column=14).value=list8[i]
    ws.cell(row=9+p+i,column=15).value=octant_names[i]

#list9 created to hold multiple dictionaries holding the key as octant and vale as count for each mod range    
list9=[]    
dic1={}
for i in range(0,p):
    dic_tmp={}
    for j in range(0,8):
        dic_tmp[list8[j]]=cp[i][j]
    dic_tmp={k:v for k, v in sorted(dic_tmp.items(), key=lambda item: item[1])} # sorting the dictionary based on values
    list9.append(dic_tmp)

#sorting the values automatically ranks the data in descending order (8 to 1)     
#list11 is a temporary list ; each time it runs and stores the sorted dictionary data as tuples of pairs of octants and counts 
#list10 contains all the temporary lists made by list11
list10=[]
for i in range(0,p):
    list11=[]
    list11=list(list9[i].items())
    list10.append(list11)
    
#writing the ranks in the excel file
for i in range(0,p):
    for j in range(0,8):
        if list10[i][j][0]==1:
            ws.cell(row=4+i,column=22).value=8-j
        if list10[i][j][0]==-1:
            ws.cell(row=4+i,column=23).value=8-j
        if list10[i][j][0]==2:
            ws.cell(row=4+i,column=24).value=8-j
        if list10[i][j][0]==-2:
            ws.cell(row=4+i,column=25).value=8-j
        if list10[i][j][0]==3:
            ws.cell(row=4+i,column=26).value=8-j
        if list10[i][j][0]==-3:
            ws.cell(row=4+i,column=27).value=8-j    
        if list10[i][j][0]==4:
            ws.cell(row=4+i,column=28).value=8-j    
        if list10[i][j][0]==-4:
            ws.cell(row=4+i,column=29).value=8-j
            
#mapping octant id with octant names
dic_on={1:"Internal outward interaction",-1:"External Outward Interaction",2:"External Ejection",-2:"Internal Ejection",3:"External Inward Interaction",-3:"Internal Inward Interaction",4:"Internal Sweep",-4:"External Sweep"}
#list12 will contain the octants ids which got rank 1 in all the ranges
list12=[]
#writing the octant id and name for the octant which got rank1 in all the ranges ; also appending the ids to list12
for i in range(0,p):
    ws.cell(row=4+i,column=30).value=list10[i][-1][0]
    ws.cell(row=4+i,column=31).value=dic_on[list10[i][-1][0]]
    list12.append(list10[i][-1][0])

#writing the count of octants with rank 1 from list12
for i in range(0,8):
        ws.cell(row=9+i+p,column=16).value=list12.count(list8[i])
        
wb.save("octant_output_ranking_excel.xlsx")

#program completed ( made on SPYDER IDE 5.3.3 with PYTHON 3.8.10 64 BIT)