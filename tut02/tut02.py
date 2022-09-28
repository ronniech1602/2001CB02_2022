from openpyxl import load_workbook
import pandas as pd
import numpy as np

wb=load_workbook("input_octant_transition_identify.xlsx")
df=pd.read_excel("input_octant_transition_identify.xlsx")
ws=wb.worksheets[0]

#creating lists to hold values of u,v,w excluding 'u','v','w' using pandas
list1=df.U
list2=df.V
list3=df.W

#calculating the summation of u,v,w values and then dividing by number of elements to get averages
totalu=0
for x in range(0,len(list1)):
    totalu=totalu+list1[x]

Uavg=totalu/len(list1)

totalv=0
for x in range(0,len(list2)):
    totalv=totalv+list2[x]

Vavg=totalv/len(list2)

totalw=0
for x in range(0,len(list3)):
    totalw=totalw+list3[x]
    
Wavg=totalw/len(list3)

#forming lists of title and value for averages , and writing them to excel file
Ua=["Uavg",Uavg]
Va=["Vavg",Vavg]
Wa=["Wavg",Wavg]
for i in range(1,3):
    ws.cell(row=i,column=5).value=Ua[i-1]

for i in range(1,3):
    ws.cell(row=i,column=6).value=Va[i-1]
    
for i in range(1,3):
    ws.cell(row=i,column=7).value=Wa[i-1]

#list4,list5,list6 to store values of u-uavg , v-vavg , w-wavg as lists
list4=[]
for i in list1:
    tmp=i-Uavg
    list4.append(tmp)
    
ws.cell(row=1,column=8).value="U'"
for i in range(2,len(list4)+2):
    ws.cell(row=i,column=8).value=list4[i-2]
    
list5=[]
for i in list2:
    tmp=i-Vavg
    list5.append(tmp)
    
ws.cell(row=1,column=9).value="V'"
for i in range(2,len(list5)+2):
    ws.cell(row=i,column=9).value=list5[i-2]
    
list6=[]
for i in list3:
    tmp=i-Wavg
    list6.append(tmp)
    
ws.cell(row=1,column=10).value="W'"
for i in range(2,len(list6)+2):
    ws.cell(row=i,column=10).value=list6[i-2]

#list7 is created to hold the octant value for u',v',w' data, as a list
ws.cell(row=1,column=11).value="Octants"
list7=[]
for i in range (0,len(list1)):
    if list4[i]>0 and list5[i]>0:
        if list6[i]>0:
            list7.append(+1)
        else:
            list7.append(-1)
        
    if list4[i]>0 and list5[i]<0:
        if list6[i]>0:
            list7.append(+4)
        else:
            list7.append(-4)
        
    if list4[i]<0 and list5[i]>0:
        if list6[i]>0:
            list7.append(+2)
        else:
            list7.append(-2)
        
    if list4[i]<0 and list5[i]<0:
        if list6[i]>0:
            list7.append(+3)
        else:
            list7.append(-3)
        

for i in range(2,len(list7)+2):
    ws.cell(row=i,column=11).value=list7[i-2]
    
#creating the table for overall counts and adding the overall counts from list7
ws.cell(row=2,column=13).value="Overall Count"
ws.cell(row=3,column=12).value="User Input"
list8=[+1,-1,+2,-2,+3,-3,+4,-4]
for i in range(0,8):
    ws.cell(row=1,column=i+14).value=list8[i]
   
ws.cell(row=2,column=14).value=list7.count(+1)
ws.cell(row=2,column=15).value=list7.count(-1)
ws.cell(row=2,column=16).value=list7.count(+2)
ws.cell(row=2,column=17).value=list7.count(-2)
ws.cell(row=2,column=18).value=list7.count(+3)
ws.cell(row=2,column=19).value=list7.count(-3)
ws.cell(row=2,column=20).value=list7.count(+4)
ws.cell(row=2,column=21).value=list7.count(-4)

#giving value to mod , which can be changed and program will behave accordingly
mod=5000
ws.cell(row=3,column=13).value="Mod"+str(mod)

#creating a partition variable p
mod_ranges=[]
p=(len(list1)//mod)+1
cl=[] #cl holds empty lists based on number of partitions
for i in range(0,p):
    l=[]
    cl.append(l)

a=0
#the empty lists in cl hold values of octants in the ranges which depend on value of mod or p
for y in range(0,p):
    for x in range(a,a+mod):
        if x<=len(list1)-1:
            cl[y].append(list7[x])
    a=a+mod

#writing the individual count in the table of variable dimensions , that depend on mod value or p value
for i in range(0,p): 
    if mod*(i+1)<=len(list1):
        ws.cell(row=4+i,column=13).value=str(mod*i)+"-"+str(mod*(i+1)-1)
    else:
        ws.cell(row=4+i,column=13).value=str(mod*i)+"-"+str(len(list1)-1)
    ws.cell(row=4+i,column=14).value=cl[i].count(+1)
    ws.cell(row=4+i,column=15).value=cl[i].count(-1)
    ws.cell(row=4+i,column=16).value=cl[i].count(+2)
    ws.cell(row=4+i,column=17).value=cl[i].count(-2)
    ws.cell(row=4+i,column=18).value=cl[i].count(+3)
    ws.cell(row=4+i,column=19).value=cl[i].count(-3)
    ws.cell(row=4+i,column=20).value=cl[i].count(+4)
    ws.cell(row=4+i,column=21).value=cl[i].count(-4)

#creating a table for overall count transition
ws.cell(row=p+7,column=13).value="Overall Count Transition"
ws.cell(row=p+8,column=14).value="To"
ws.cell(row=p+10,column=12).value="From"
ws.cell(row=p+9,column=13).value="Count"
for i in range(0,8):
    ws.cell(row=p+9,column=14+i).value=list8[i]
    ws.cell(row=p+10+i,column=13).value=list8[i]
 
#calculating the overall transition count and storing it as lists A1-A8
a1=b1=c1=d1=e1=f1=g1=h1=0
a2=b2=c2=d2=e2=f2=g2=h2=0
a3=b3=c3=d3=e3=f3=g3=h3=0
a4=b4=c4=d4=e4=f4=g4=h4=0
a5=b5=c5=d5=e5=f5=g5=h5=0
a6=b6=c6=d6=e6=f6=g6=h6=0
a7=b7=c7=d7=e7=f7=g7=h7=0
a8=b8=c8=d8=e8=f8=g8=h8=0

for i in range(0,len(list7)-1):
    if list7[i]==1:
        if list7[i+1]==1:
            a1+=1
        elif list7[i+1]==-1:
            b1+=1
        elif list7[i+1]==2:
            c1+=1
        elif list7[i+1]==-2:
            d1+=1
        elif list7[i+1]==3:
            e1+=1
        elif list7[i+1]==-3:
            f1+=1
        elif list7[i+1]==4:
            g1+=1
        elif list7[i+1]==-4:
            h1+=1
    if list7[i]==-1:
        if list7[i+1]==1:
            a2+=1
        elif list7[i+1]==-1:
            b2+=1
        elif list7[i+1]==2:
            c2+=1
        elif list7[i+1]==-2:
            d2+=1
        elif list7[i+1]==3:
            e2+=1
        elif list7[i+1]==-3:
            f2+=1
        elif list7[i+1]==4:
            g2+=1
        elif list7[i+1]==-4:
            h2+=1
    if list7[i]==2:
        if list7[i+1]==1:
            a3+=1
        elif list7[i+1]==-1:
            b3+=1
        elif list7[i+1]==2:
            c3+=1
        elif list7[i+1]==-2:
            d3+=1
        elif list7[i+1]==3:
            e3+=1
        elif list7[i+1]==-3:
            f3+=1
        elif list7[i+1]==4:
            g3+=1
        elif list7[i+1]==-4:
            h3+=1
    if list7[i]==-2:
        if list7[i+1]==1:
            a4+=1
        elif list7[i+1]==-1:
            b4+=1
        elif list7[i+1]==2:
            c4+=1
        elif list7[i+1]==-2:
            d4+=1
        elif list7[i+1]==3:
            e4+=1
        elif list7[i+1]==-3:
            f4+=1
        elif list7[i+1]==4:
            g4+=1
        elif list7[i+1]==-4:
            h4+=1
    if list7[i]==3:
        if list7[i+1]==1:
            a5+=1
        elif list7[i+1]==-1:
            b5+=1
        elif list7[i+1]==2:
            c5+=1
        elif list7[i+1]==-2:
            d5+=1
        elif list7[i+1]==3:
            e5+=1
        elif list7[i+1]==-3:
            f5+=1
        elif list7[i+1]==4:
            g5+=1
        elif list7[i+1]==-4:
            h5+=1
    if list7[i]==-3:
        if list7[i+1]==1:
            a6+=1
        elif list7[i+1]==-1:
            b6+=1
        elif list7[i+1]==2:
            c6+=1
        elif list7[i+1]==-2:
            d6+=1
        elif list7[i+1]==3:
            e6+=1
        elif list7[i+1]==-3:
            f6+=1
        elif list7[i+1]==4:
            g6+=1
        elif list7[i+1]==-4:
            h6+=1
    if list7[i]==4:
        if list7[i+1]==1:
            a7+=1
        elif list7[i+1]==-1:
            b7+=1
        elif list7[i+1]==2:
            c7+=1
        elif list7[i+1]==-2:
            d7+=1
        elif list7[i+1]==3:
            e7+=1
        elif list7[i+1]==-3:
            f7+=1
        elif list7[i+1]==4:
            g7+=1
        elif list7[i+1]==-4:
            h7+=1
    if list7[i]==-4:
        if list7[i+1]==1:
            a8+=1
        elif list7[i+1]==-1:
            b8+=1
        elif list7[i+1]==2:
            c8+=1
        elif list7[i+1]==-2:
            d8+=1
        elif list7[i+1]==3:
            e8+=1
        elif list7[i+1]==-3:
            f8+=1
        elif list7[i+1]==4:
            g8+=1
        elif list7[i+1]==-4:
            h8+=1    

A1=[a1,b1,c1,d1,e1,f1,g1,h1]
A2=[a2,b2,c2,d2,e2,f2,g2,h2]
A3=[a3,b3,c3,d3,e3,f3,g3,h3]
A4=[a4,b4,c4,d4,e4,f4,g4,h4]
A5=[a5,b5,c5,d5,e5,f5,g5,h5]
A6=[a6,b6,c6,d6,e6,f6,g6,h6]
A7=[a7,b7,c7,d7,e7,f7,g7,h7]
A8=[a8,b8,c8,d8,e8,f8,g8,h8]

#writing the overall transition count values to excel
for i in range(0,8):
    ws.cell(row=p+10,column=14+i).value=A1[i]
    ws.cell(row=p+11,column=14+i).value=A2[i]
    ws.cell(row=p+12,column=14+i).value=A3[i]
    ws.cell(row=p+13,column=14+i).value=A4[i]
    ws.cell(row=p+14,column=14+i).value=A5[i]
    ws.cell(row=p+15,column=14+i).value=A6[i]
    ws.cell(row=p+16,column=14+i).value=A7[i]
    ws.cell(row=p+17,column=14+i).value=A8[i]

#calculating the mod transition count (individual transition count) and storing the values as lists A10-A80
for x in range(0,p):
    if x<p:
        ws.cell(row=p+21+13*x,column=13).value="Mod Transition Count"
        if mod*(x+1)<=len(list1):
            ws.cell(row=p+22+13*x,column=13).value=str(mod*x)+"-"+str(mod*(x+1)-1)
        else:
            ws.cell(row=p+22+13*x,column=13).value=str(mod*x)+"-"+str(len(list1)-1)
        ws.cell(row=p+22+13*x,column=14).value="To"
        ws.cell(row=p+24+13*x,column=12).value="From"
        ws.cell(row=p+23+13*x,column=13).value="Count"
        for i in range(0,8):
            ws.cell(row=p+23+13*x,column=14+i).value=list8[i]
            ws.cell(row=p+24+13*x+i,column=13).value=list8[i]
        a10=b10=c10=d10=e10=f10=g10=h10=0
        a20=b20=c20=d20=e20=f20=g20=h20=0
        a30=b30=c30=d30=e30=f30=g30=h30=0
        a40=b40=c40=d40=e40=f40=g40=h40=0
        a50=b50=c50=d50=e50=f50=g50=h50=0
        a60=b60=c60=d60=e60=f60=g60=h60=0
        a70=b70=c70=d70=e70=f70=g70=h70=0
        a80=b80=c80=d80=e80=f80=g80=h80=0

        for i in range(0,len(cl[x])-1):
            if cl[x][i]==1:
                if cl[x][i+1]==1:
                    a10+=1
                elif cl[x][i+1]==-1:
                    b10+=1
                elif cl[x][i+1]==2:
                    c10+=1
                elif cl[x][i+1]==-2:
                    d10+=1
                elif cl[x][i+1]==3:
                    e10+=1
                elif cl[x][i+1]==-3:
                    f10+=1
                elif cl[x][i+1]==4:
                    g10+=1
                elif cl[x][i+1]==-4:
                    h10+=1
            if cl[x][i]==-1:
                if cl[x][i+1]==1:
                    a20+=1
                elif cl[x][i+1]==-1:
                    b20+=1
                elif cl[x][i+1]==2:
                    c20+=1
                elif cl[x][i+1]==-2:
                    d20+=1
                elif cl[x][i+1]==3:
                    e20+=1
                elif cl[x][i+1]==-3:
                    f20+=1
                elif cl[x][i+1]==4:
                    g20+=1
                elif cl[x][i+1]==-4:
                    h20+=1
            if cl[x][i]==2:
                if cl[x][i+1]==1:
                    a30+=1
                elif cl[x][i+1]==-1:
                    b30+=1
                elif cl[x][i+1]==2:
                    c30+=1
                elif cl[x][i+1]==-2:
                    d30+=1
                elif cl[x][i+1]==3:
                    e30+=1
                elif cl[x][i+1]==-3:
                    f30+=1
                elif cl[x][i+1]==4:
                    g30+=1
                elif cl[x][i+1]==-4:
                    h30+=1
            if cl[x][i]==-2:
                if cl[x][i+1]==1:
                    a40+=1
                elif cl[x][i+1]==-1:
                    b40+=1
                elif cl[x][i+1]==2:
                    c40+=1
                elif cl[x][i+1]==-2:
                    d40+=1
                elif cl[x][i+1]==3:
                    e40+=1
                elif cl[x][i+1]==-3:
                    f40+=1
                elif cl[x][i+1]==4:
                    g40+=1
                elif cl[x][i+1]==-4:
                    h40+=1
            if cl[x][i]==3:
                if cl[x][i+1]==1:
                    a50+=1
                elif cl[x][i+1]==-1:
                    b50+=1
                elif cl[x][i+1]==2:
                    c50+=1
                elif cl[x][i+1]==-2:
                    d50+=1
                elif cl[x][i+1]==3:
                    e50+=1
                elif cl[x][i+1]==-3:
                    f50+=1
                elif cl[x][i+1]==4:
                    g50+=1
                elif cl[x][i+1]==-4:
                    h50+=1
            if cl[x][i]==-3:
                if cl[x][i+1]==1:
                    a60+=1
                elif cl[x][i+1]==-1:
                    b60+=1
                elif cl[x][i+1]==2:
                    c60+=1
                elif cl[x][i+1]==-2:
                    d60+=1
                elif cl[x][i+1]==3:
                    e60+=1
                elif cl[x][i+1]==-3:
                    f60+=1
                elif cl[x][i+1]==4:
                    g60+=1
                elif cl[x][i+1]==-4:
                    h60+=1
            if cl[x][i]==4:
                if cl[x][i+1]==1:
                    a70+=1
                elif cl[x][i+1]==-1:
                    b70+=1
                elif cl[x][i+1]==2:
                    c70+=1
                elif cl[x][i+1]==-2:
                    d70+=1
                elif cl[x][i+1]==3:
                    e70+=1
                elif cl[x][i+1]==-3:
                    f70+=1
                elif cl[x][i+1]==4:
                    g70+=1
                elif cl[x][i+1]==-4:
                    h70+=1
            if cl[x][i]==-4:
                if cl[x][i+1]==1:
                    a80+=1
                elif cl[x][i+1]==-1:
                    b80+=1
                elif cl[x][i+1]==2:
                    c80+=1
                elif cl[x][i+1]==-2:
                    d80+=1
                elif cl[x][i+1]==3:
                    e80+=1
                elif cl[x][i+1]==-3:
                    f80+=1
                elif cl[x][i+1]==4:
                    g80+=1
                elif cl[x][i+1]==-4:
                    h80+=1    

        A10=[a10,b10,c10,d10,e10,f10,g10,h10]
        A20=[a20,b20,c20,d20,e20,f20,g20,h20]
        A30=[a30,b30,c30,d30,e30,f30,g30,h30]
        A40=[a40,b40,c40,d40,e40,f40,g40,h40]
        A50=[a50,b50,c50,d50,e50,f50,g50,h50]
        A60=[a60,b60,c60,d60,e60,f60,g60,h60]
        A70=[a70,b70,c70,d70,e70,f70,g70,h70]
        A80=[a80,b80,c80,d80,e80,f80,g80,h80]
        
        #writing the individual transitions counts to excel
        for i in range(0,8):
            ws.cell(row=p+24+13*x,column=14+i).value=A10[i]
            ws.cell(row=p+25+13*x,column=14+i).value=A20[i]
            ws.cell(row=p+26+13*x,column=14+i).value=A30[i]
            ws.cell(row=p+27+13*x,column=14+i).value=A40[i]
            ws.cell(row=p+28+13*x,column=14+i).value=A50[i]
            ws.cell(row=p+29+13*x,column=14+i).value=A60[i]
            ws.cell(row=p+30+13*x,column=14+i).value=A70[i]
            ws.cell(row=p+31+13*x,column=14+i).value=A80[i]
            
wb.save('output_octant_transition_identify.xlsx') #saving the output file
    
#program completed ( made on SPYDER IDE 5.3.3 with PYTHON 3.8.10 64 BIT)


