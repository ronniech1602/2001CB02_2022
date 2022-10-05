from openpyxl import load_workbook
import pandas as pd
import numpy as np

wb=load_workbook("input_octant_longest_subsequence.xlsx")
df=pd.read_excel("input_octant_longest_subsequence.xlsx")
ws=wb.worksheets[0]

#reading excel data as panda dataframe and storing it as lists
list1=df.U
list2=df.V
list3=df.W

#calculating the summation of values of u,v,w and their averages
totalu=0
for x in range(0,len(list1)):
    totalu=totalu+list1[x]

Uavg=totalu/len(list1)

totalv=0
for x in range(0,len(list2)):
    totalv=totalv+list2[x]

Vavg=totalv/len(list2)

totalw=0
for x in range(0,len(list3)):
    totalw=totalw+list3[x]
    
Wavg=totalw/len(list3)

#writing the values of averages in excel 
Ua=["Uavg",Uavg]
Va=["Vavg",Vavg]
Wa=["Wavg",Wavg]
for i in range(1,3):
    ws.cell(row=i,column=5).value=Ua[i-1]

for i in range(1,3):
    ws.cell(row=i,column=6).value=Va[i-1]
    
for i in range(1,3):
    ws.cell(row=i,column=7).value=Wa[i-1]

#calculating the diffences u-uavg,v-vavg,w-wavg ;storing as lists and writing them into excel
list4=[]
for i in list1:
    tmp=i-Uavg
    list4.append(tmp)
    
ws.cell(row=1,column=8).value="U'"
for i in range(2,len(list4)+2):
    ws.cell(row=i,column=8).value=list4[i-2]
    
list5=[]
for i in list2:
    tmp=i-Vavg
    list5.append(tmp)
    
ws.cell(row=1,column=9).value="V'"
for i in range(2,len(list5)+2):
    ws.cell(row=i,column=9).value=list5[i-2]
    
list6=[]
for i in list3:
    tmp=i-Wavg
    list6.append(tmp)
    
ws.cell(row=1,column=10).value="W'"
for i in range(2,len(list6)+2):
    ws.cell(row=i,column=10).value=list6[i-2]

#calculating the octant numbers and storing them as list7
ws.cell(row=1,column=11).value="Octants"
list7=[]
for i in range (0,len(list1)):
    if list4[i]>0 and list5[i]>0:
        if list6[i]>0:
            list7.append(+1)
        else:
            list7.append(-1)
        
    if list4[i]>0 and list5[i]<0:
        if list6[i]>0:
            list7.append(+4)
        else:
            list7.append(-4)
        
    if list4[i]<0 and list5[i]>0:
        if list6[i]>0:
            list7.append(+2)
        else:
            list7.append(-2)
        
    if list4[i]<0 and list5[i]<0:
        if list6[i]>0:
            list7.append(+3)
        else:
            list7.append(-3)
        
#writing list7 to excel
for i in range(2,len(list7)+2):
    ws.cell(row=i,column=11).value=list7[i-2]

#pos is a list of lists , with each list containing the indices of a particular octant number from list7
octants=[1,-1,2,-2,3,-3,4,-4]
pos=[]
for i in range(0,8):
    l=[]
    pos.append(l)

for x in range(0,len(list7)):
    if list7[x]==1:
        pos[0].append(x)
    elif list7[x]==-1:
        pos[1].append(x)
    elif list7[x]==2:
        pos[2].append(x)
    elif list7[x]==-2:
        pos[3].append(x)
    elif list7[x]==3:
        pos[4].append(x)
    elif list7[x]==-3:
        pos[5].append(x)
    elif list7[x]==4:
        pos[6].append(x)
    elif list7[x]==-4:
        pos[7].append(x)

#forming the sublists of consecutive integers from lists in pos        
x1=[]
sx1=[]
c1=-1
for n in pos[0]:
    if c1+1!=n:           
        if sx1:              
            x1.append(sx1)
            sx1=[]
    sx1.append(n)
    c1=n
if sx1:
    x1.append(sx1)

v1=[len(x1[i]) for i in range(0,len(x1))] #v1 contains lengths of lists formed in x1
y1=max(v1) #longest list length in x1 , which is the longest length for which the corresponding octant appears in list7
z1=v1.count(max(v1)) #count of longest list length in x1

x1=[]
sx1=[]
c1=-1
for n in pos[1]:
    if c1+1!=n:           
        if sx1:              
            x1.append(sx1)
            sx1=[]
    sx1.append(n)
    c1=n
if sx1:
    x1.append(sx1)

v2=[len(x1[i]) for i in range(0,len(x1))]
y2=max(v2)
z2=v2.count(max(v2))

x1=[]
sx1=[]
c1=-1
for n in pos[2]:
    if c1+1!=n:           
        if sx1:              
            x1.append(sx1)
            sx1=[]
    sx1.append(n)
    c1=n
if sx1:
    x1.append(sx1)

v3=[len(x1[i]) for i in range(0,len(x1))]
y3=max(v3)
z3=v3.count(max(v3))

x1=[]
sx1=[]
c1=-1
for n in pos[3]:
    if c1+1!=n:           
        if sx1:              
            x1.append(sx1)
            sx1=[]
    sx1.append(n)
    c1=n
if sx1:
    x1.append(sx1)

v4=[len(x1[i]) for i in range(0,len(x1))]
y4=max(v4)
z4=v4.count(max(v4))

x1=[]
sx1=[]
c1=-1
for n in pos[4]:
    if c1+1!=n:           
        if sx1:              
            x1.append(sx1)
            sx1=[]
    sx1.append(n)
    c1=n
if sx1:
    x1.append(sx1)

v5=[len(x1[i]) for i in range(0,len(x1))]
y5=max(v5)
z5=v5.count(max(v5))

x1=[]
sx1=[]
c1=-1
for n in pos[5]:
    if c1+1!=n:           
        if sx1:              
            x1.append(sx1)
            sx1=[]
    sx1.append(n)
    c1=n
if sx1:
    x1.append(sx1)

v6=[len(x1[i]) for i in range(0,len(x1))]
y6=max(v6)
z6=v6.count(max(v6))

x1=[]
sx1=[]
c1=-1
for n in pos[6]:
    if c1+1!=n:           
        if sx1:              
            x1.append(sx1)
            sx1=[]
    sx1.append(n)
    c1=n
if sx1:
    x1.append(sx1)

v7=[len(x1[i]) for i in range(0,len(x1))]
y7=max(v7)
z7=v7.count(max(v7))

x1=[]
sx1=[]
c1=-1
for n in pos[7]:
    if c1+1!=n:           
        if sx1:              
            x1.append(sx1)
            sx1=[]
    sx1.append(n)
    c1=n
if sx1:
    x1.append(sx1)

v8=[len(x1[i]) for i in range(0,len(x1))]
y8=max(v8)
z8=v8.count(max(v8))

#Y has all the longest length of appearances of all octants and Z has its counts
Y=[y1,y2,y3,y4,y5,y6,y7,y8]
Z=[z1,z2,z3,z4,z5,z6,z7,z8]

#writing the longest subsequence length and count to excel
ws.cell(row=2,column=13).value="Count"
ws.cell(row=2,column=14).value="Longest Subsequence Length"
ws.cell(row=2,column=15).value="Count"
for i in range(0,8):
    ws.cell(row=3+i,column=13).value=octants[i]
    ws.cell(row=3+i,column=14).value=Y[i]
    ws.cell(row=3+i,column=15).value=Z[i]

wb.save("output_octant_longest_subsequence.xlsx")

#program completed ( made on SPYDER IDE 5.3.3 with PYTHON 3.8.10 64 BIT)
        
