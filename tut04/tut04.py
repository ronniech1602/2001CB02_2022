from openpyxl import load_workbook
import pandas as pd
import numpy as np

wb=load_workbook("input_octant_longest_subsequence_with_range.xlsx")
df=pd.read_excel("input_octant_longest_subsequence_with_range.xlsx")
ws=wb.worksheets[0]

#reading excel data as panda dataframe and storing it as lists
list1=df.U
list2=df.V
list3=df.W

#calculating the summation of values of u,v,w and their averages
totalu=0
for x in range(0,len(list1)):
    totalu=totalu+list1[x]

Uavg=totalu/len(list1)

totalv=0
for x in range(0,len(list2)):
    totalv=totalv+list2[x]

Vavg=totalv/len(list2)

totalw=0
for x in range(0,len(list3)):
    totalw=totalw+list3[x]
    
Wavg=totalw/len(list3)

#writing the values of averages in excel 
Ua=["Uavg",Uavg]
Va=["Vavg",Vavg]
Wa=["Wavg",Wavg]
for i in range(1,3):
    ws.cell(row=i,column=5).value=Ua[i-1]

for i in range(1,3):
    ws.cell(row=i,column=6).value=Va[i-1]
    
for i in range(1,3):
    ws.cell(row=i,column=7).value=Wa[i-1]

#calculating the diffences u-uavg,v-vavg,w-wavg ;storing as lists and writing them into excel
list4=[]
for i in list1:
    tmp=i-Uavg
    list4.append(tmp)
    
ws.cell(row=1,column=8).value="U'"
for i in range(2,len(list4)+2):
    ws.cell(row=i,column=8).value=list4[i-2]
    
list5=[]
for i in list2:
    tmp=i-Vavg
    list5.append(tmp)
    
ws.cell(row=1,column=9).value="V'"
for i in range(2,len(list5)+2):
    ws.cell(row=i,column=9).value=list5[i-2]
    
list6=[]
for i in list3:
    tmp=i-Wavg
    list6.append(tmp)
    
ws.cell(row=1,column=10).value="W'"
for i in range(2,len(list6)+2):
    ws.cell(row=i,column=10).value=list6[i-2]

#calculating the octant numbers and storing them as list7
ws.cell(row=1,column=11).value="Octants"
list7=[]
for i in range (0,len(list1)):
    if list4[i]>0 and list5[i]>0:
        if list6[i]>0:
            list7.append(+1)
        else:
            list7.append(-1)
        
    if list4[i]>0 and list5[i]<0:
        if list6[i]>0:
            list7.append(+4)
        else:
            list7.append(-4)
        
    if list4[i]<0 and list5[i]>0:
        if list6[i]>0:
            list7.append(+2)
        else:
            list7.append(-2)
        
    if list4[i]<0 and list5[i]<0:
        if list6[i]>0:
            list7.append(+3)
        else:
            list7.append(-3)
        
#writing list7 to excel
for i in range(2,len(list7)+2):
    ws.cell(row=i,column=11).value=list7[i-2]

#pos is a list of lists , with each list containing the indices of a particular octant number from list7
octants=[1,-1,2,-2,3,-3,4,-4]
pos=[]
for i in range(0,8):
    l=[]
    pos.append(l)

for x in range(0,len(list7)):
    if list7[x]==1:
        pos[0].append(x)
    elif list7[x]==-1:
        pos[1].append(x)
    elif list7[x]==2:
        pos[2].append(x)
    elif list7[x]==-2:
        pos[3].append(x)
    elif list7[x]==3:
        pos[4].append(x)
    elif list7[x]==-3:
        pos[5].append(x)
    elif list7[x]==4:
        pos[6].append(x)
    elif list7[x]==-4:
        pos[7].append(x)

#forming the sublists of consecutive integers from lists in pos        
x1=[]
sx1=[]
c1=-1
for n in pos[0]:
    if c1+1!=n:           
        if sx1:              
            x1.append(sx1)
            sx1=[]
    sx1.append(n)
    c1=n
if sx1:
    x1.append(sx1)

v1=[len(x1[i]) for i in range(0,len(x1))] #v1 contains lengths of lists formed in x1
y1=max(v1) #longest list length in x1 , which is the longest length for which the corresponding octant appears in list7
z1=v1.count(max(v1)) #count of longest list length in x1

x2=[]
sx2=[]
c2=-1
for n in pos[1]:
    if c2+1!=n:           
        if sx2:              
            x2.append(sx2)
            sx2=[]
    sx2.append(n)
    c2=n
if sx2:
    x2.append(sx2)

v2=[len(x2[i]) for i in range(0,len(x2))]
y2=max(v2)
z2=v2.count(max(v2))

x3=[]
sx3=[]
c3=-1
for n in pos[2]:
    if c3+1!=n:           
        if sx3:              
            x3.append(sx3)
            sx3=[]
    sx3.append(n)
    c3=n
if sx3:
    x3.append(sx3)

v3=[len(x3[i]) for i in range(0,len(x3))]
y3=max(v3)
z3=v3.count(max(v3))

x4=[]
sx4=[]
c4=-1
for n in pos[3]:
    if c4+1!=n:           
        if sx4:              
            x4.append(sx4)
            sx4=[]
    sx4.append(n)
    c4=n
if sx4:
    x4.append(sx4)

v4=[len(x4[i]) for i in range(0,len(x4))]
y4=max(v4)
z4=v4.count(max(v4))

x5=[]
sx5=[]
c5=-1
for n in pos[4]:
    if c5+1!=n:           
        if sx5:              
            x5.append(sx5)
            sx5=[]
    sx5.append(n)
    c5=n
if sx5:
    x5.append(sx5)

v5=[len(x5[i]) for i in range(0,len(x5))]
y5=max(v5)
z5=v5.count(max(v5))

x6=[]
sx6=[]
c6=-1
for n in pos[5]:
    if c6+1!=n:           
        if sx6:              
            x6.append(sx6)
            sx6=[]
    sx6.append(n)
    c6=n
if sx6:
    x6.append(sx6)

v6=[len(x6[i]) for i in range(0,len(x6))]
y6=max(v6)
z6=v6.count(max(v6))

x7=[]
sx7=[]
c7=-1
for n in pos[6]:
    if c7+1!=n:           
        if sx7:              
            x7.append(sx7)
            sx7=[]
    sx7.append(n)
    c7=n
if sx7:
    x7.append(sx7)

v7=[len(x7[i]) for i in range(0,len(x7))]
y7=max(v7)
z7=v7.count(max(v7))

x8=[]
sx8=[]
c8=-1
for n in pos[7]:
    if c8+1!=n:           
        if sx8:              
            x1.append(sx8)
            sx8=[]
    sx8.append(n)
    c8=n
if sx8:
    x8.append(sx8)

v8=[len(x8[i]) for i in range(0,len(x8))]
y8=max(v8)
z8=v8.count(max(v8))

#Y has all the longest length of appearances of all octants and Z has its counts
Y=[y1,y2,y3,y4,y5,y6,y7,y8]
Z=[z1,z2,z3,z4,z5,z6,z7,z8]

#writing the longest subsequence length and count to excel
ws.cell(row=2,column=13).value="Count"
ws.cell(row=2,column=14).value="Longest Subsequence Length"
ws.cell(row=2,column=15).value="Count"
for i in range(0,8):
    ws.cell(row=3+i,column=13).value=octants[i]
    ws.cell(row=3+i,column=14).value=Y[i]
    ws.cell(row=3+i,column=15).value=Z[i]

#creating a list of time values using pandas and writing the header of skeleton    
listT=df.Time 
ws.cell(row=2,column=17).value="Count"
ws.cell(row=2,column=18).value="Longest Subsequence Length"
ws.cell(row=2,column=19).value="Count"
#creating the skeleton for mentioning time ranges with counts of longest subsequences
o=0
for i in range(0,8):
    ws.cell(row=3+i+o,column=17).value=octants[i]
    ws.cell(row=3+i+o,column=18).value=Y[i]
    ws.cell(row=3+i+o,column=19).value=Z[i]
    ws.cell(row=4+i+o,column=17).value="Time"
    ws.cell(row=4+i+o,column=18).value="From"
    ws.cell(row=4+i+o,column=19).value="To"
    o=o+Z[i]+1

#creating lists to hold the longest subsequence indices of each octant    
x_1=[]    
for x in x1:
    if len(x)==y1:
        x_1.append(x)
x_2=[]    
for x in x2:
    if len(x)==y2:
        x_2.append(x)
x_3=[]    
for x in x3:
    if len(x)==y3:
        x_3.append(x)
x_4=[]    
for x in x4:
    if len(x)==y4:
        x_4.append(x)
x_5=[]    
for x in x5:
    if len(x)==y5:
        x_5.append(x)
x_6=[]    
for x in x6:
    if len(x)==y6:
        x_6.append(x)
x_7=[]    
for x in x7:
    if len(x)==y7:
        x_7.append(x)
x_8=[]    
for x in x8:
    if len(x)==y8:
        x_8.append(x)

#writing the values of time ranges to excel using the above created lists x_1 to x_8
X=[x_1,x_2,x_3,x_4,x_5,x_6,x_7,x_8]
o=0        
for i in range(0,8):
    for n in range(0,Z[i]):
        ws.cell(row=5+i+o+n,column=18).value=listT[X[i][n][0]]
        ws.cell(row=5+i+o+n,column=19).value=listT[X[i][n][-1]]
    o=o+Z[i]+1
    
wb.save("output_octant_longest_subsequence_with_range.xlsx")

#program completed ( made on SPYDER IDE 5.3.3 with PYTHON 3.8.10 64 BIT)